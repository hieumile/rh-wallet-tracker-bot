"""
Exports the GMGN-based pipeline's per-wallet aggregates (from
scoring/wallet_aggregator.py) to an Excel workbook.

One tab, "Wallet Summary", one row per wallet, with the numbers the pipeline
already computed from GMGN — total buy/sell (USD), net, trade counts, profit on
this token, and GMGN's wallet-level winrate / 7d profit / pnl ratio. Values are
written directly (not via SUMIFS) because each wallet is already a single
aggregated row here, not a stream of individual trade events.

Updated to support Approach B: merging runs into a master spreadsheet, keeping
historical transactions, and grouping PnL by wallet + token.
"""

import os
import logging
from datetime import datetime, timezone
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scoring.wallet_aggregator import WalletAggregate, insider_signals

logger = logging.getLogger(__name__)

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial")


def _style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _ts_to_dt(ts: int | None):
    if ts is None:
        return None
    return datetime.fromtimestamp(ts, tz=timezone.utc).replace(tzinfo=None)


def _num(value):
    """Parse GMGN's string/number fields to float, or None."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_existing_transactions(file_path: str) -> list[dict]:
    if not os.path.exists(file_path):
        return []
    try:
        wb = openpyxl.load_workbook(file_path)
        if "Transactions" not in wb.sheetnames:
            return []
        ws = wb["Transactions"]
        headers = [cell.value for cell in ws[1]]
        has_token_address = "Token Address" in headers
        
        events = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if has_token_address:
                if len(row) < 9:
                    continue
                dt, wallet, token_address, side, symbol, amount, value, gas, tx_hash = row[:9]
            else:
                if len(row) < 8:
                    continue
                dt, wallet, side, symbol, amount, value, gas, tx_hash = row[:8]
                token_address = ""
            
            if not wallet or not tx_hash:
                continue
            ts = int(dt.replace(tzinfo=timezone.utc).timestamp()) if isinstance(dt, datetime) else None
            events.append({
                "wallet": wallet,
                "timestamp": ts,
                "event_type": side,
                "token": {"address": token_address or "", "symbol": symbol},
                "token_amount": amount,
                "cost_usd": value,
                "gas_usd": gas,
                "tx_hash": tx_hash
            })
        return events
    except Exception as e:
        logger.warning("Could not load existing transactions from %s: %s", file_path, e)
        return []



def export_transactions(events: list[dict], output_path: str, raw_events: list[dict] = None):
    """
    Workbook from raw wallet events:
      - "Transactions": one row per buy/sell, newest first. Merges with existing transactions.
      - "Wallet PnL": one row per wallet per token combining all its buys and sells, with PnL.
      - "Raw Transactions": optional tab containing raw unfiltered logs with their filter status.
    """
    if not events:
        raise ValueError("No transactions to export — nothing was ingested.")

    # Load existing transaction history to merge
    existing = _load_existing_transactions(output_path)
    
    # Avoid duplicate events by using a unique key: (tx_hash, wallet, token_address)
    merged_map = {}
    for ev in existing:
        tx = ev.get("tx_hash")
        w = ev.get("wallet")
        t = (ev.get("token") or {}).get("address") or ""
        key = (tx, w, t)
        merged_map[key] = ev

    for ev in events:
        tx = ev.get("tx_hash")
        w = ev.get("wallet")
        token_info = ev.get("token") or {}
        t = token_info.get("address") or ""
        key = (tx, w, t)
        side = (ev.get("event_type") or ev.get("type") or "").lower()
        merged_map[key] = {
            "wallet": w,
            "timestamp": ev.get("timestamp"),
            "event_type": side,
            "token": {"address": t, "symbol": token_info.get("symbol") or "TOKEN"},
            "token_amount": ev.get("token_amount"),
            "cost_usd": ev.get("cost_usd"),
            "gas_usd": ev.get("gas_usd"),
            "tx_hash": tx
        }

    merged_events = list(merged_map.values())
    rows = sorted(merged_events, key=lambda e: int(e.get("timestamp") or 0), reverse=True)

    wb = Workbook()

    # ---------- Transactions tab ----------
    ws = wb.active
    ws.title = "Transactions"
    headers = [
        "Time (UTC)", "Wallet", "Token Address", "Side", "Symbol",
        "Token Amount", "Value (USD)", "Gas (USD)", "Tx Hash",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for e in rows:
        side = (e.get("event_type") or e.get("type") or "").lower()
        token = e.get("token") or {}
        ts = e.get("timestamp")
        ws.append([
            _ts_to_dt(int(ts)) if ts else None,
            e.get("wallet"),
            token.get("address"),
            side,
            token.get("symbol"),
            _num(e.get("token_amount")),
            _num(e.get("cost_usd")),
            _num(e.get("gas_usd")),
            e.get("tx_hash"),
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=r, column=6).number_format = "#,##0.########"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"
        ws.cell(row=r, column=8).number_format = "$#,##0.00"

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
    ws.column_dimensions["B"].width = 44  # wallet
    ws.column_dimensions["C"].width = 44  # token address
    ws.column_dimensions["I"].width = 68  # tx hash



    # ---------- Raw Transactions tab ----------
    if raw_events:
        ws3 = wb.create_sheet("Raw Transactions")
        headers3 = [
            "Time (UTC)", "Wallet", "Token Address", "Side", "Symbol",
            "Token Amount", "Value (USD)", "Tx Hash", "Status / Filter Reason"
        ]
        ws3.append(headers3)
        _style_header_row(ws3, 1, len(headers3))
        
        sorted_raw = sorted(raw_events, key=lambda e: int(e.get("timestamp") or 0), reverse=True)
        
        for e in sorted_raw:
            token = e.get("token") or {}
            ts = e.get("timestamp")
            ws3.append([
                _ts_to_dt(int(ts)) if ts else None,
                e.get("wallet"),
                token.get("address"),
                (e.get("event_type") or "").upper(),
                token.get("symbol"),
                _num(e.get("token_amount")),
                _num(e.get("cost_usd")),
                e.get("tx_hash"),
                e.get("status")
            ])
            r = ws3.max_row
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=r, column=c).font = BODY_FONT
            ws3.cell(row=r, column=1).number_format = "yyyy-mm-dd hh:mm:ss"
            ws3.cell(row=r, column=6).number_format = "#,##0.########"
            ws3.cell(row=r, column=7).number_format = "$#,##0.00"
            
        for i, header in enumerate(headers3, start=1):
            ws3.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
        ws3.column_dimensions["B"].width = 44  # wallet
        ws3.column_dimensions["C"].width = 44  # token address
        ws3.column_dimensions["H"].width = 68  # tx hash
        ws3.column_dimensions["I"].width = 28  # status / reason

    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    wb.save(output_path)


def _load_existing_wallet_summary(file_path: str) -> dict[tuple[str, str], dict]:
    if not os.path.exists(file_path):
        return {}
    try:
        wb = openpyxl.load_workbook(file_path)
        if "Wallet Summary" not in wb.sheetnames:
            return {}
        ws = wb["Wallet Summary"]
        
        # Read header row to map columns dynamically
        header_row = [str(cell.value).strip().lower() for cell in ws[1]]
        
        def get_val(row, name, default=None):
            try:
                idx = header_row.index(name.lower())
                return row[idx] if idx < len(row) else default
            except ValueError:
                return default
                
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
                
            wallet = get_val(row, "Wallet")
            token = get_val(row, "Token")
            if not wallet or not token:
                continue
                
            key = (str(wallet).lower(), str(token).lower())
            data[key] = {
                "wallet": wallet,
                "signals": get_val(row, "Insider Signals", ""),
                "tags": get_val(row, "Tags", ""),
                "token": token,
                "buy_usd": get_val(row, "Total Buy (USD)"),
                "sell_usd": get_val(row, "Total Sell (USD)"),
                "net_usd": get_val(row, "Net (USD)"),
                "buy_count": get_val(row, "Buy Count"),
                "sell_count": get_val(row, "Sell Count"),
                "profit": get_val(row, "Profit (USD)"),
                "profit_change": get_val(row, "Profit Change %"),
                "winrate": get_val(row, "Winrate %"),
                "wallet_profit": get_val(row, "Wallet Profit (period)"),
                "wallet_volume_usd": get_val(row, "Wallet Volume (USD)"),
                "wallet_pnl_ratio": get_val(row, "Wallet PnL Ratio"),
                "wallet_profit_factor": get_val(row, "Wallet Profit Factor"),
                "wallet_sharpe_ratio": get_val(row, "Wallet Sharpe Ratio"),
                "wallet_tx_count": get_val(row, "Wallet Tx Count"),
                "wallet_max_drawdown_ratio": get_val(row, "Wallet Max Drawdown %"),
                "first_trade": get_val(row, "First Trade (UTC)"),
                "last_trade": get_val(row, "Last Trade (UTC)"),
            }
        return data
    except Exception as e:
        logger.warning("Could not load existing wallet summary from %s: %s", file_path, e)
        return {}


def export_wallet_report(
    aggregates: list[WalletAggregate],
    output_path: str,
    token_label: str = "TOKEN",
):
    """
    Write per-wallet aggregates to `output_path`. Ranks by per-token profit
    (falling back to net USD when GMGN didn't report a profit for the token).
    Missing values are left blank rather than written as a misleading 0.
    Merges with existing reports to compile a master list.
    """
    if not aggregates:
        raise ValueError("No wallet aggregates to export — nothing to write.")

    # Load existing summaries to merge
    merged_data = _load_existing_wallet_summary(output_path)

    # Add/update with new aggregates
    for a in aggregates:
        token_addr = (a.token_address or token_label).lower()
        key = (a.wallet.lower(), token_addr)
        
        profit_change_pct = a.profit_change * 100 if a.profit_change is not None else None
        winrate_pct = a.winrate * 100 if a.winrate is not None else None
        drawdown_pct = a.wallet_max_drawdown_ratio * 100 if a.wallet_max_drawdown_ratio is not None else None
        
        merged_data[key] = {
            "wallet": a.wallet,
            "signals": ", ".join(insider_signals(a)),
            "tags": ",".join(a.tags),
            "token": a.token_address or token_label,
            "buy_usd": a.total_buy_usd,
            "sell_usd": a.total_sell_usd,
            "net_usd": a.net_usd,
            "buy_count": a.buy_count,
            "sell_count": a.sell_count,
            "profit": a.realized_profit_usd,
            "profit_change": profit_change_pct,
            "winrate": winrate_pct,
            "wallet_profit": a.wallet_realized_profit,
            "wallet_volume_usd": a.wallet_volume_usd,
            "wallet_pnl_ratio": a.wallet_pnl_ratio,
            "wallet_profit_factor": a.wallet_profit_factor,
            "wallet_sharpe_ratio": a.wallet_sharpe_ratio,
            "wallet_tx_count": a.wallet_tx_count,
            "wallet_max_drawdown_ratio": drawdown_pct,
            "first_trade": _ts_to_dt(a.first_ts),
            "last_trade": _ts_to_dt(a.last_ts),
        }

    # Sort merged rows by profit (falling back to net USD)
    def sort_key(item):
        val = item["profit"] if item["profit"] is not None else item["net_usd"]
        return (val is not None, val or 0.0)

    rows = sorted(merged_data.values(), key=sort_key, reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Wallet Summary"

    headers = [
        "Wallet", "Insider Signals", "Tags", "Token",
        "Total Buy (USD)", "Total Sell (USD)", "Net (USD)",
        "Buy Count", "Sell Count",
        "Profit (USD)", "Profit Change %",
        "Winrate %", "Wallet Profit (period)", "Wallet Volume (USD)", "Wallet PnL Ratio",
        "Wallet Profit Factor", "Wallet Sharpe Ratio", "Wallet Tx Count",
        "Wallet Max Drawdown %", "First Trade (UTC)", "Last Trade (UTC)",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for r_data in rows:
        ws.append([
            r_data["wallet"],
            r_data["signals"],
            r_data["tags"],
            r_data["token"],
            r_data["buy_usd"],
            r_data["sell_usd"],
            r_data["net_usd"],
            r_data["buy_count"],
            r_data["sell_count"],
            r_data["profit"],
            r_data["profit_change"],
            r_data["winrate"],
            r_data["wallet_profit"],
            r_data.get("wallet_volume_usd"),
            r_data["wallet_pnl_ratio"],
            r_data.get("wallet_profit_factor"),
            r_data.get("wallet_sharpe_ratio"),
            r_data.get("wallet_tx_count"),
            r_data.get("wallet_max_drawdown_ratio"),
            r_data["first_trade"],
            r_data["last_trade"],
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
        for c in (5, 6, 7, 10, 13, 14):   # USD columns
            ws.cell(row=r, column=c).number_format = "$#,##0"
        for c in (11, 12, 19):            # percentage columns
            ws.cell(row=r, column=c).number_format = "0.0"
        ws.cell(row=r, column=15).number_format = "0.000"        # pnl ratio
        ws.cell(row=r, column=16).number_format = "0.00"         # profit factor
        ws.cell(row=r, column=17).number_format = "0.00"         # Sharpe ratio
        ws.cell(row=r, column=18).number_format = "#,##0"        # tx count
        for c in (20, 21):                # timestamps
            ws.cell(row=r, column=c).number_format = "yyyy-mm-dd hh:mm:ss"

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
    ws.column_dimensions["A"].width = 44  # wallet address
    ws.column_dimensions["B"].width = 34  # insider signals
    ws.column_dimensions["D"].width = 44  # token address

    note_row = len(rows) + 3
    ws.cell(
        row=note_row, column=1,
        value=(
            "Note: 'Profit', 'Total Buy/Sell' and Buy/Sell counts are for THIS token. In "
            "windowed mode (--from/--to) they are computed from wallet activity within the "
            "date range, and Profit = sum over in-window sells of (proceeds - cost basis); "
            "otherwise they are GMGN's all-time token_top_traders figures. 'Winrate', "
            "'Wallet 7d Profit' and 'Wallet PnL Ratio' are GMGN's cross-token 7d stats for "
            "the wallet. 'Insider Signals' are rule-based flags (GMGN rat_trader/sniper/"
            "smart_money tags, high winrate, profitable). Blank cells mean GMGN did not "
            "report that value (not zero)."
        ),
    ).font = Font(name="Arial", italic=True, size=9, color="808080")

    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    wb.save(output_path)


def export_signals_report(signals: list[dict], co_investments: list[dict], output_path: str):
    """
    Export signal generator results to an Excel workbook:
      - "Transaction Signals": lists recent buys/sells by followed wallets.
      - "Co-Investments": lists tokens bought by 2 or more followed wallets.
    """
    wb = Workbook()

    # ---------- Transaction Signals tab ----------
    ws = wb.active
    ws.title = "Transaction Signals"
    headers = [
        "Time (UTC)", "Wallet", "Score", "Side", "Symbol", "Token Name",
        "Token Address", "Token Amount", "Estimated Value (USD)", "Tx Hash",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for s in signals:
        ts = s.get("timestamp")
        ws.append([
            _ts_to_dt(int(ts)) if ts else None,
            s.get("wallet"),
            s.get("score"),
            s.get("side"),
            s.get("symbol"),
            s.get("name"),
            s.get("token_address"),
            s.get("amount"),
            s.get("estimated_value_usd"),
            s.get("tx_hash"),
        ])
        r = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=col_idx).font = BODY_FONT
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=r, column=3).number_format = "0.0"
        ws.cell(row=r, column=8).number_format = "#,##0.########"
        ws.cell(row=r, column=9).number_format = "$#,##0.00"

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
    ws.column_dimensions["B"].width = 44  # wallet
    ws.column_dimensions["G"].width = 44  # token address
    ws.column_dimensions["J"].width = 68  # tx hash

    # ---------- Co-Investments tab ----------
    ws2 = wb.create_sheet("Co-Investments")
    headers2 = [
        "Symbol", "Token Name", "Token Address", "Buyers Count", 
        "Current Price (USD)", "Liquidity (USD)", "Buyers (Wallet [Score, Amount, Time])"
    ]
    ws2.append(headers2)
    _style_header_row(ws2, 1, len(headers2))

    for c in co_investments:
        buyers_str = "; ".join([
            f"{b['wallet'][:6]}...{b['wallet'][-4:]} [Score: {b['score']}, Amount: {b['amount']:.0f}]"
            for b in c["buyers"]
        ])
        ws2.append([
            c["symbol"],
            c["name"],
            c["token_address"],
            c["buyers_count"],
            c.get("price_usd"),
            c.get("liquidity_usd"),
            buyers_str
        ])
        r = ws2.max_row
        for col_idx in range(1, len(headers2) + 1):
            ws2.cell(row=r, column=col_idx).font = BODY_FONT
        ws2.cell(row=r, column=4).number_format = "#,##0"
        ws2.cell(row=r, column=5).number_format = "$#,##0.00######"
        ws2.cell(row=r, column=6).number_format = "$#,##0.00"

    for i, header in enumerate(headers2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
    ws2.column_dimensions["C"].width = 44  # token address
    ws2.column_dimensions["G"].width = 80  # buyers string

    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    wb.save(output_path)


def export_watchlist(watchlist_map: dict[str, dict], output_path: str):
    """
    Export the scored wallet watchlist to an Excel sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    
    headers = [
        "Wallet", "Score", "Win Rate", "Realized Profit (USD)", "Volume (USD)", "PnL Ratio",
        "Token Count", "Moonshot Count", "Max Drawdown %", "Profit Factor", "Sharpe Ratio", "Tx Count",
        "Tags", "Insider Signals", "Seed Tokens", "Last Updated"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    
    # Sort by score descending
    sorted_entries = sorted(watchlist_map.values(), key=lambda e: float(e.get("score") or 0.0), reverse=True)
    
    for e in sorted_entries:
        winrate = e.get("winrate")
        if winrate is not None:
            winrate = float(winrate)
            
        dd_ratio = e.get("max_drawdown_ratio")
        if dd_ratio is not None:
            dd_ratio = float(dd_ratio)
            
        ws.append([
            e.get("wallet"),
            _num(e.get("score")),
            winrate,
            _num(e.get("realized_profit_usd")),
            _num(e.get("volume_usd")),
            _num(e.get("pnl_ratio")),
            e.get("token_num"),
            e.get("moonshot_count"),
            dd_ratio,
            _num(e.get("profit_factor")),
            _num(e.get("sharpe_ratio")),
            e.get("tx_count"),
            ", ".join(e.get("tags") or []),
            ", ".join(e.get("insider_signals") or []),
            ", ".join(e.get("seed_tokens") or []),
            e.get("updated_at")
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=2).number_format = "0.0"
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "$#,##0.00"
        ws.cell(row=r, column=5).number_format = "$#,##0.00"
        ws.cell(row=r, column=6).number_format = "0.00"
        ws.cell(row=r, column=9).number_format = "0.0%"
        ws.cell(row=r, column=10).number_format = "0.00"
        ws.cell(row=r, column=11).number_format = "0.00"
        ws.cell(row=r, column=12).number_format = "#,##0"
        
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 2)
    ws.column_dimensions["A"].width = 44  # wallet
    ws.column_dimensions["M"].width = 30  # tags
    ws.column_dimensions["N"].width = 40  # signals
    ws.column_dimensions["O"].width = 30  # seed tokens
    
    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    wb.save(output_path)
